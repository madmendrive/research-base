"""Excel model builder — generates financial model from extracted filing data.

Builds a workbook with company-specific line items (not a fixed template).
Uses the exact labels from SEC filings to preserve fidelity.
"""

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"
CONFIG_PATH = PROJECT_ROOT / "config" / "companies.json"
MODELS_DIR = PROJECT_ROOT / "models"

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
BLUE = Font(color="0000FF")
BLUE_B = Font(color="0000FF", bold=True)
BLACK = Font(color="000000")
BLACK_B = Font(color="000000", bold=True)
GREEN = Font(color="00B050")
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

NUM_FMT = "#,##0"
EPS_FMT = "0.00"
PCT_FMT = "0.0%"


def _cl(col):
    return get_column_letter(col)


# ---------------------------------------------------------------------------
# Period helpers
# ---------------------------------------------------------------------------

_FY_RE = re.compile(r"^FY(\d{4})$")
_Q_RE = re.compile(r"^Q([1-4])\s+(\d{4})$")
_YTD_PATTERNS = re.compile(
    r"^(Nine Months|9M|Six Months|6M|H[12]|Three Months Ended)", re.IGNORECASE
)


def _parse_period(pk):
    """Parse a period key. Returns (type, year, quarter) or None.
    type: 'annual', 'quarterly', or None (skip)."""
    m = _FY_RE.match(pk)
    if m:
        return ("annual", int(m.group(1)), None)
    m = _Q_RE.match(pk)
    if m:
        return ("quarterly", int(m.group(2)), int(m.group(1)))
    return None


def _classify_is_periods(period_keys):
    """Classify IS period keys into annual years and quarterly (year, q) tuples.
    Skips YTD/interim periods."""
    annual_years = []
    quarterly = []
    for pk in period_keys:
        parsed = _parse_period(pk)
        if not parsed:
            continue
        typ, year, q = parsed
        if typ == "annual":
            annual_years.append(year)
        elif typ == "quarterly":
            quarterly.append((year, q))
    annual_years = sorted(set(annual_years))
    quarterly = sorted(set(quarterly))
    return annual_years, quarterly


# ---------------------------------------------------------------------------
# Column map
# ---------------------------------------------------------------------------

def _build_col_map(annual_years, quarterly):
    """Build column index mapping. Returns dict."""
    # Annual: C onwards
    annual_cols = {}
    for i, yr in enumerate(annual_years[-9:]):  # max 9 years
        annual_cols[yr] = 3 + i

    spacer_col = 3 + len(annual_cols)

    # Quarterly: all quarters for annual years + any orphan quarters
    all_q = []
    for yr in sorted(annual_cols.keys()):
        for q in [1, 2, 3, 4]:
            all_q.append((yr, q))
    for yr, q in quarterly:
        if yr not in annual_cols:
            all_q.append((yr, q))
    # Deduplicate and sort
    seen = set()
    ordered_q = []
    for item in all_q:
        if item not in seen:
            seen.add(item)
            ordered_q.append(item)
    ordered_q.sort()

    quarterly_cols = {}
    for i, (yr, q) in enumerate(ordered_q):
        quarterly_cols[(yr, q)] = spacer_col + 1 + i

    return {
        "annual": annual_cols,
        "quarterly": quarterly_cols,
        "spacer_col": spacer_col,
        "ordered_q": ordered_q,
    }


def _all_data_cols(col_map):
    cols = list(col_map["annual"].values()) + list(col_map["quarterly"].values())
    return sorted(cols)


def _period_key_for_col(col_map, col):
    """Return the period key string for a given column index."""
    for yr, c in col_map["annual"].items():
        if c == col:
            return f"FY{yr}"
    for (yr, q), c in col_map["quarterly"].items():
        if c == col:
            return f"Q{q} {yr}"
    return None


# ---------------------------------------------------------------------------
# Cell writing helpers
# ---------------------------------------------------------------------------

def _val(ws, row, col, value, font=BLUE, fmt=NUM_FMT):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.number_format = fmt


def _formula(ws, row, col, formula, font=BLACK, fmt=NUM_FMT):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = font
    cell.number_format = fmt


def _label(ws, row, text, bold=False, indent=False):
    cell = ws.cell(row=row, column=2, value=text)
    if bold:
        cell.font = Font(bold=True)
    if indent:
        cell.alignment = Alignment(indent=1)


def _section_header(ws, row, text, col_range=None):
    cell = ws.cell(row=row, column=2, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    if col_range:
        for c in col_range:
            ws.cell(row=row, column=c).fill = HEADER_FILL


def _write_period_headers(ws, row, col_map, annual_years):
    for yr, col in col_map["annual"].items():
        cell = ws.cell(row=row, column=col, value=yr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for (yr, q), col in col_map["quarterly"].items():
        cell = ws.cell(row=row, column=col, value=f"Q{q}-{str(yr)[-2:]}")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Detect if an IS item is per-share
# ---------------------------------------------------------------------------

def _is_per_share(label):
    lower = label.lower()
    return "per share" in lower or "earnings per" in lower or lower.startswith("eps")


def _is_shares_item(label):
    lower = label.lower()
    return "shares" in lower and ("outstanding" in lower or "used" in lower
                                   or "weighted" in lower)


# ---------------------------------------------------------------------------
# IS section builder
# ---------------------------------------------------------------------------

def _write_is_section(ws, data, col_map, annual_years, quarterly):
    """Write income statement with exact company line items.
    Returns (next_row, revenue_row, dict of label->row for analytics)."""
    is_data = data.get("income_statement", {})
    line_items = is_data.get("line_items", [])

    if not line_items:
        return 6, None, {}

    header_row = 5
    _section_header(ws, header_row, "INCOME STATEMENT",
                    col_range=_all_data_cols(col_map))
    _write_period_headers(ws, header_row, col_map, annual_years)

    row = 6
    revenue_row = None
    label_to_row = {}
    last_subtotal_row = None

    for item in line_items:
        lbl = item.get("label", "")
        item_type = item.get("type", "input")
        values = item.get("values", {})
        formula_desc = item.get("formula_description", "")
        is_ps = _is_per_share(lbl)
        is_shares = _is_shares_item(lbl)
        fmt = EPS_FMT if (is_ps or item_type == "per_share") else NUM_FMT

        is_subtotal = item_type == "subtotal"
        _label(ws, row, lbl, bold=is_subtotal)
        label_to_row[lbl] = row

        if revenue_row is None and not is_subtotal:
            revenue_row = row  # First input line is revenue

        # Determine which rows this subtotal references (for Excel formulas)
        formula_rows = None
        if is_subtotal and formula_desc:
            formula_rows = _parse_formula_desc(formula_desc, label_to_row, row)

        # Write annual columns
        for yr, col in col_map["annual"].items():
            pk = f"FY{yr}"
            val = values.get(pk)
            if is_subtotal and formula_rows:
                f_str = _build_subtotal_formula(formula_rows, col, row,
                                                 last_subtotal_row)
                _formula(ws, row, col, f_str, font=BLACK_B, fmt=fmt)
            elif val is not None:
                _val(ws, row, col, val, font=BLUE, fmt=fmt)

        # Write quarterly columns
        for yr, q in col_map["ordered_q"]:
            col = col_map["quarterly"][(yr, q)]
            pk = f"Q{q} {yr}"

            if q == 4 and yr in col_map["annual"]:
                # Q4 = Annual - SUM(Q1:Q3)
                ann_c = _cl(col_map["annual"][yr])
                q1_key = (yr, 1)
                q3_key = (yr, 3)
                if q1_key in col_map["quarterly"] and q3_key in col_map["quarterly"]:
                    q1_c = _cl(col_map["quarterly"][q1_key])
                    q3_c = _cl(col_map["quarterly"][q3_key])
                    q4_formula = f"={ann_c}{row}-SUM({q1_c}{row}:{q3_c}{row})"
                    q4_font = BLUE if not is_subtotal else BLACK_B
                    _formula(ws, row, col, q4_formula, font=q4_font, fmt=fmt)
                continue

            if is_subtotal and formula_rows:
                f_str = _build_subtotal_formula(formula_rows, col, row,
                                                 last_subtotal_row)
                _formula(ws, row, col, f_str, font=BLACK_B, fmt=fmt)
            else:
                val = values.get(pk)
                if val is not None:
                    _val(ws, row, col, val, font=BLUE, fmt=fmt)

        if is_subtotal:
            last_subtotal_row = row
        row += 1

    return row + 1, revenue_row, label_to_row


def _parse_formula_desc(desc, label_to_row, current_row):
    """Parse a formula_description to determine Excel formula components.
    Returns a list of (operator, row_num) tuples, or None if can't parse.

    Falls back to a range-sum of all input rows since the last subtotal."""
    if not desc:
        return None

    desc_lower = desc.lower()

    # "Sum of ..." pattern — sum contiguous block above
    if "sum" in desc_lower and "of" in desc_lower:
        return [("sum_range", None)]

    # Try "A - B" or "A + B" patterns
    # Split on + or - that are surrounded by spaces (to avoid splitting
    # inside labels like "Other income/(expense)")
    parts = re.split(r'\s+([+\-])\s+', desc)

    if len(parts) >= 3:
        result = []
        sign = "+"
        for part in parts:
            part = part.strip()
            if part in ("+", "-"):
                sign = part
                continue
            matched_row = _find_label_row(part, label_to_row)
            if matched_row:
                result.append((sign, matched_row))

        if len(result) >= 2:
            return result

    # Fallback: try the full string as "A + B" with comma separation
    if "," in desc:
        result = []
        for part in desc.split(","):
            part = part.strip()
            matched_row = _find_label_row(part, label_to_row)
            if matched_row:
                result.append(("+", matched_row))
        if len(result) >= 2:
            return result

    return None


def _find_label_row(text, label_to_row):
    """Find the row number for a label. Exact match only — no fuzzy."""
    text_clean = text.strip().lower()
    # Exact match
    for lbl, row in label_to_row.items():
        if lbl.lower() == text_clean:
            return row
    # Case-insensitive match stripping punctuation
    text_stripped = re.sub(r'[^\w\s]', '', text_clean)
    for lbl, row in label_to_row.items():
        lbl_stripped = re.sub(r'[^\w\s]', '', lbl.lower())
        if text_stripped == lbl_stripped:
            return row
    return None


def _build_subtotal_formula(formula_rows, col, current_row, last_subtotal_row=None):
    """Build an Excel formula string from parsed formula components."""
    c = _cl(col)

    if not formula_rows:
        return None

    # Handle sum_range — sum all rows from last subtotal+1 to current-1
    if formula_rows[0][0] == "sum_range":
        start = (last_subtotal_row + 1) if last_subtotal_row else (current_row - 1)
        end = current_row - 1
        if start > end:
            start = end
        return f"=SUM({c}{start}:{c}{end})"

    # Build A + B - C style formula
    parts = []
    for i, (sign, row_num) in enumerate(formula_rows):
        ref = f"{c}{row_num}"
        if i == 0:
            if sign == "-":
                parts.append(f"-{ref}")
            else:
                parts.append(ref)
        else:
            parts.append(f"{sign}{ref}")

    return "=" + "".join(parts)


# ---------------------------------------------------------------------------
# Y/Y, Q/Q, % of sales sections
# ---------------------------------------------------------------------------

def _find_key_rows(label_to_row):
    """Find rows for key analytical items based on common label patterns."""
    key_items = []
    patterns = [
        (r"(?:net\s+)?(?:sales|revenue)", "Revenue"),
        (r"gross\s+(?:margin|profit)", "Gross profit"),
        (r"(?:total\s+)?operating\s+(?:income|expenses)", "Operating income"),
        (r"(?:net\s+)?income", "Net income"),
    ]
    for pattern, display_name in patterns:
        for lbl, row in label_to_row.items():
            if re.search(pattern, lbl, re.IGNORECASE):
                key_items.append((display_name, lbl, row))
                break
    return key_items


def _write_yy_section(ws, row, col_map, annual_years, label_to_row):
    """Write Y/Y growth. Returns next row."""
    key_items = _find_key_rows(label_to_row)
    if not key_items:
        return row

    _section_header(ws, row, "Y/Y", col_range=_all_data_cols(col_map))
    _write_period_headers(ws, row, col_map, annual_years)
    row += 1

    sorted_annual = sorted(col_map["annual"].items())
    ordered_q = col_map["ordered_q"]
    q_cols = col_map["quarterly"]

    for display_name, lbl, src_row in key_items:
        _label(ws, row, display_name)

        # Annual Y/Y
        for i in range(1, len(sorted_annual)):
            curr_col = sorted_annual[i][1]
            prev_col = sorted_annual[i - 1][1]
            f = f"={_cl(curr_col)}{src_row}/{_cl(prev_col)}{src_row}-1"
            _formula(ws, row, curr_col, f, fmt=PCT_FMT)

        # Quarterly Y/Y
        for yr, q in ordered_q:
            prev = (yr - 1, q)
            if prev in q_cols:
                curr_c = q_cols[(yr, q)]
                prev_c = q_cols[prev]
                f = f"={_cl(curr_c)}{src_row}/{_cl(prev_c)}{src_row}-1"
                _formula(ws, row, curr_c, f, fmt=PCT_FMT)

        row += 1

    return row + 1


def _write_qq_section(ws, row, col_map, annual_years, label_to_row):
    """Write Q/Q growth. Returns next row."""
    key_items = _find_key_rows(label_to_row)
    if not key_items:
        return row

    _section_header(ws, row, "Q/Q", col_range=_all_data_cols(col_map))
    _write_period_headers(ws, row, col_map, annual_years)
    row += 1

    ordered_q = col_map["ordered_q"]
    q_cols = col_map["quarterly"]

    for display_name, lbl, src_row in key_items:
        _label(ws, row, display_name)

        for idx in range(1, len(ordered_q)):
            curr = ordered_q[idx]
            prev = ordered_q[idx - 1]
            curr_c = q_cols[curr]
            prev_c = q_cols[prev]
            f = f"={_cl(curr_c)}{src_row}/{_cl(prev_c)}{src_row}-1"
            _formula(ws, row, curr_c, f, fmt=PCT_FMT)

        row += 1

    return row + 1


def _write_pct_section(ws, row, col_map, annual_years, label_to_row, revenue_row):
    """Write % of sales section. Returns next row."""
    if not revenue_row:
        return row

    _section_header(ws, row, "As a % of sales", col_range=_all_data_cols(col_map))
    _write_period_headers(ws, row, col_map, annual_years)
    row += 1

    # Build list of items to show as % of revenue
    pct_items = []
    for lbl, src_row in label_to_row.items():
        if src_row == revenue_row:
            continue
        pct_items.append((lbl, src_row))

    # Limit to key items to keep it manageable
    key_labels = []
    patterns = [
        r"cost\s+of\s+(?:sales|goods|revenue)",
        r"gross\s+(?:margin|profit)",
        r"research|R&D",
        r"selling|SG&A|general\s+and\s+admin",
        r"(?:total\s+)?operating\s+(?:expense|income)",
        r"(?:profit|income)\s+before\s+(?:tax|provision)",
        r"(?:net\s+)?income",
    ]
    for pattern in patterns:
        for lbl, src_row in pct_items:
            if re.search(pattern, lbl, re.IGNORECASE):
                key_labels.append((lbl, src_row))
                break

    all_cols = _all_data_cols(col_map)
    for lbl, src_row in key_labels:
        _label(ws, row, lbl)
        for col in all_cols:
            c = _cl(col)
            f = f"={c}{src_row}/{c}${revenue_row}"
            _formula(ws, row, col, f, fmt=PCT_FMT)
        row += 1

    return row + 1


# ---------------------------------------------------------------------------
# Balance Sheet section (annual only)
# ---------------------------------------------------------------------------

def _bs_val_for_year(values, yr, fy_to_bs_pk):
    """Get BS value for a fiscal year, trying various key formats."""
    pk_fy = f"FY{yr}"
    if pk_fy in values:
        return values[pk_fy]
    if yr in fy_to_bs_pk:
        return values.get(fy_to_bs_pk[yr])
    # Fallback: any key containing the year
    for pk, v in values.items():
        if str(yr) in pk:
            return v
    return None


def _write_bs_section(ws, data, col_map, annual_years, start_row):
    """Write balance sheet with exact company line items. Annual only.
    Groups items by section, writes subtotals as SUM formulas.
    Returns (next_row, label_to_row_bs)."""
    bs_data = data.get("balance_sheet", {})
    line_items = bs_data.get("line_items", [])

    if not line_items:
        return start_row, {}

    row = start_row
    ann_cols = sorted(col_map["annual"].values())

    _section_header(ws, row, "BALANCE SHEET", col_range=ann_cols)
    for yr, col in col_map["annual"].items():
        cell = ws.cell(row=row, column=col)
        cell.value = f"={_cl(col)}5"
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    label_to_row_bs = {}

    # Map BS period_keys to annual columns
    bs_period_keys = bs_data.get("period_keys", [])
    fy_to_bs_pk = {}
    for yr in annual_years:
        for pk in bs_period_keys:
            if str(yr) in pk:
                fy_to_bs_pk[yr] = pk
                break

    # Group items by section, preserving order within each section
    section_order = ["current_assets", "non_current_assets",
                     "current_liabilities", "non_current_liabilities", "equity"]
    section_items = {s: [] for s in section_order}
    # Items that are "total" across sections (Total assets, Total liabilities, etc.)
    cross_section_totals = []

    for item in line_items:
        sec = item.get("section", "")
        lbl = item.get("label", "").lower()
        # Detect cross-section totals
        if item.get("type") == "subtotal" and (
            "total assets" in lbl or
            "total liabilities and" in lbl or
            ("total liabilities" in lbl and "equity" not in lbl
             and "shareholders" not in lbl and sec == "non_current_liabilities")
        ):
            cross_section_totals.append(item)
            continue
        if sec in section_items:
            section_items[sec].append(item)
        else:
            # Unknown section — skip duplicates, add to equity as fallback
            section_items.setdefault("equity", []).append(item)

    # Deduplicate items by label (keep first occurrence with values)
    for sec in section_items:
        seen = {}
        deduped = []
        for item in section_items[sec]:
            lbl = item.get("label", "")
            # Truncate long labels for dedup
            lbl_short = lbl[:60]
            if lbl_short in seen:
                # Merge values into existing item
                existing = deduped[seen[lbl_short]]
                existing.setdefault("values", {}).update(item.get("values", {}))
            else:
                seen[lbl_short] = len(deduped)
                deduped.append(item)
        section_items[sec] = deduped

    section_labels = {
        "current_assets": "Current assets",
        "non_current_assets": "Non-current assets",
        "current_liabilities": "Current liabilities",
        "non_current_liabilities": "Non-current liabilities",
        "equity": "Equity",
    }

    subtotal_rows = {}  # section_name -> row

    for sec in section_order:
        items = section_items.get(sec, [])
        if not items:
            continue

        if subtotal_rows:
            row += 1  # blank row between sections

        _label(ws, row, section_labels.get(sec, sec), bold=True)
        row += 1
        section_start = row

        # Separate input items from subtotals
        inputs = [it for it in items if it.get("type") != "subtotal"]
        subtotals = [it for it in items if it.get("type") == "subtotal"]

        for item in inputs:
            lbl = item.get("label", "")
            # Truncate very long labels (common stock descriptions)
            if len(lbl) > 70:
                lbl = lbl[:67] + "..."
            _label(ws, row, lbl)
            label_to_row_bs[item.get("label", "")] = row
            values = item.get("values", {})
            for yr, col in col_map["annual"].items():
                val = _bs_val_for_year(values, yr, fy_to_bs_pk)
                if val is not None:
                    _val(ws, row, col, val)
            row += 1

        # Write section subtotal
        for item in subtotals:
            lbl = item.get("label", "")
            _label(ws, row, lbl, bold=True)
            label_to_row_bs[lbl] = row
            for col in ann_cols:
                c = _cl(col)
                _formula(ws, row, col, f"=SUM({c}{section_start}:{c}{row - 1})",
                         font=BLACK_B)
            subtotal_rows[sec] = row
            row += 1

    # Cross-section totals
    for item in cross_section_totals:
        lbl = item.get("label", "").lower()
        label_full = item.get("label", "")
        _label(ws, row, label_full, bold=True)
        label_to_row_bs[label_full] = row

        if "total assets" in lbl:
            # = total current assets + total non-current assets
            ca_row = subtotal_rows.get("current_assets")
            nca_row = subtotal_rows.get("non_current_assets")
            if ca_row and nca_row:
                for col in ann_cols:
                    c = _cl(col)
                    _formula(ws, row, col, f"={c}{ca_row}+{c}{nca_row}",
                             font=BLACK_B)
        elif "total liabilities and" in lbl:
            # = total liabilities + total equity
            tl_row = subtotal_rows.get("_total_liabilities")
            te_row = subtotal_rows.get("equity")
            if tl_row and te_row:
                for col in ann_cols:
                    c = _cl(col)
                    _formula(ws, row, col, f"={c}{tl_row}+{c}{te_row}",
                             font=BLACK_B)
        elif "total liabilities" in lbl:
            # = total current liabilities + total non-current liabilities
            cl_row = subtotal_rows.get("current_liabilities")
            ncl_row = subtotal_rows.get("non_current_liabilities")
            if cl_row and ncl_row:
                for col in ann_cols:
                    c = _cl(col)
                    _formula(ws, row, col, f"={c}{cl_row}+{c}{ncl_row}",
                             font=BLACK_B)
            subtotal_rows["_total_liabilities"] = row

        row += 1

    return row + 1, label_to_row_bs


# ---------------------------------------------------------------------------
# Cash Flow section (annual only)
# ---------------------------------------------------------------------------

def _write_cf_section(ws, data, col_map, annual_years, start_row):
    """Write cash flow statement with exact company line items. Annual only.
    Returns next row."""
    cf_data = data.get("cash_flow", {})
    line_items = cf_data.get("line_items", [])

    if not line_items:
        return start_row

    row = start_row
    ann_cols = sorted(col_map["annual"].values())

    _section_header(ws, row, "CASH FLOW STATEMENT", col_range=ann_cols)
    for yr, col in col_map["annual"].items():
        cell = ws.cell(row=row, column=col)
        cell.value = f"={_cl(col)}5"
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    current_section = None
    section_start = None
    section_labels = {
        "operating": "CASH FLOWS FROM OPERATING ACTIVITIES",
        "investing": "CASH FLOWS FROM INVESTING ACTIVITIES",
        "financing": "CASH FLOWS FROM FINANCING ACTIVITIES",
    }
    label_to_row_cf = {}

    for item in line_items:
        lbl = item.get("label", "")
        section = item.get("section", "")
        item_type = item.get("type", "input")
        values = item.get("values", {})
        is_subtotal = item_type == "subtotal"

        # Section change
        if section and section != current_section:
            if current_section and section_start:
                pass  # subtotal already written
            current_section = section
            sec_label = section_labels.get(section, section.replace("_", " ").title())
            _label(ws, row, sec_label, bold=True)
            row += 1
            section_start = row

        indent = not is_subtotal and section == "operating"
        _label(ws, row, lbl, bold=is_subtotal, indent=indent)
        label_to_row_cf[lbl] = row

        if is_subtotal and section_start:
            for col in ann_cols:
                c = _cl(col)
                _formula(ws, row, col, f"=SUM({c}{section_start}:{c}{row - 1})",
                         font=BLACK_B)
            section_start = row + 1
        else:
            for yr, col in col_map["annual"].items():
                val = None
                pk_fy = f"FY{yr}"
                if pk_fy in values:
                    val = values[pk_fy]
                else:
                    for pk, v in values.items():
                        if str(yr) in pk:
                            val = v
                            break
                if val is not None:
                    _val(ws, row, col, val)

        row += 1

    return row + 1


# ---------------------------------------------------------------------------
# Drivers tab (Tab 2)
# ---------------------------------------------------------------------------

def _build_drivers_sheet(wb, company_name, data, col_map, annual_years, fs_name,
                         revenue_row):
    """Build the Drivers tab with segment data."""
    ws = wb.create_sheet(title=f"{company_name} Drivers"[:31])

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 45
    for col in _all_data_cols(col_map):
        ws.column_dimensions[_cl(col)].width = 14
    ws.column_dimensions[_cl(col_map["spacer_col"])].width = 2

    cell = ws.cell(row=3, column=2, value=company_name)
    cell.font = Font(bold=True)

    row = 5
    _section_header(ws, row, "INCOME STATEMENT ASSUMPTIONS",
                    col_range=_all_data_cols(col_map))
    _write_period_headers(ws, row, col_map, annual_years)
    row += 1

    _label(ws, row, "REVENUE", bold=True)
    row += 1

    # Revenue linked from FS tab
    rev_row_drivers = row
    _label(ws, row, "Revenue")
    if revenue_row:
        for col in _all_data_cols(col_map):
            c = _cl(col)
            _formula(ws, row, col, f"='{fs_name}'!{c}{revenue_row}", font=GREEN)
    row += 1

    # Segment data
    seg = data.get("segments", {})
    seg_names = seg.get("segment_names", [])
    seg_revenue = seg.get("data", {}).get("revenue", {})

    if seg_names:
        seg_start_row = row
        for seg_name in seg_names:
            _label(ws, row, seg_name)
            seg_vals = seg_revenue.get("values", {}).get(seg_name, {})
            # Write actual segment values
            for yr, col in col_map["annual"].items():
                pk = f"FY{yr}"
                val = seg_vals.get(pk)
                if val is not None:
                    _val(ws, row, col, val)
            for (yr, q), col in col_map["quarterly"].items():
                pk = f"Q{q} {yr}"
                val = seg_vals.get(pk)
                if val is not None:
                    _val(ws, row, col, val)
            row += 1

        row += 1

        # Revenue mix
        _label(ws, row, "Revenue mix", bold=True)
        row += 1
        for i, seg_name in enumerate(seg_names):
            _label(ws, row, seg_name)
            seg_data_row = seg_start_row + i
            for col in _all_data_cols(col_map):
                c = _cl(col)
                _formula(ws, row, col,
                         f"={c}{seg_data_row}/{c}{rev_row_drivers}",
                         fmt=PCT_FMT)
            row += 1

        row += 1

        # Q/Q growth for segments
        _label(ws, row, "Q/Q", bold=True)
        _write_period_headers(ws, row, col_map, annual_years)
        row += 1
        ordered_q = col_map["ordered_q"]
        q_cols = col_map["quarterly"]
        for i, seg_name in enumerate(seg_names):
            _label(ws, row, seg_name)
            seg_data_row = seg_start_row + i
            for idx in range(1, len(ordered_q)):
                curr_c = q_cols[ordered_q[idx]]
                prev_c = q_cols[ordered_q[idx - 1]]
                _formula(ws, row, curr_c,
                         f"={_cl(curr_c)}{seg_data_row}/{_cl(prev_c)}{seg_data_row}-1",
                         fmt=PCT_FMT)
            row += 1

        row += 1

        # Y/Y growth for segments
        _label(ws, row, "Y/Y", bold=True)
        _write_period_headers(ws, row, col_map, annual_years)
        row += 1
        for i, seg_name in enumerate(seg_names):
            _label(ws, row, seg_name)
            seg_data_row = seg_start_row + i
            # Annual Y/Y
            sorted_annual = sorted(col_map["annual"].items())
            for j in range(1, len(sorted_annual)):
                curr_col = sorted_annual[j][1]
                prev_col = sorted_annual[j - 1][1]
                _formula(ws, row, curr_col,
                         f"={_cl(curr_col)}{seg_data_row}/{_cl(prev_col)}{seg_data_row}-1",
                         fmt=PCT_FMT)
            # Quarterly Y/Y
            for yr, q in ordered_q:
                prev = (yr - 1, q)
                if prev in q_cols:
                    curr_c = q_cols[(yr, q)]
                    prev_c = q_cols[prev]
                    _formula(ws, row, curr_c,
                             f"={_cl(curr_c)}{seg_data_row}/{_cl(prev_c)}{seg_data_row}-1",
                             fmt=PCT_FMT)
            row += 1

        # Operating income by segment
        seg_oi = seg.get("data", {}).get("operating_income", {})
        if seg_oi.get("values"):
            row += 1
            _label(ws, row, "OPERATING INCOME BY SEGMENT", bold=True)
            row += 1
            oi_start = row
            for seg_name in seg_names:
                _label(ws, row, seg_name)
                oi_vals = seg_oi.get("values", {}).get(seg_name, {})
                for yr, col in col_map["annual"].items():
                    pk = f"FY{yr}"
                    val = oi_vals.get(pk)
                    if val is not None:
                        _val(ws, row, col, val)
                for (yr, q), col in col_map["quarterly"].items():
                    pk = f"Q{q} {yr}"
                    val = oi_vals.get(pk)
                    if val is not None:
                        _val(ws, row, col, val)
                row += 1

            # Segment margins
            row += 1
            _label(ws, row, "Segment operating margin", bold=True)
            row += 1
            for i, seg_name in enumerate(seg_names):
                _label(ws, row, seg_name)
                seg_rev_row = seg_start_row + i
                seg_oi_row = oi_start + i
                for col in _all_data_cols(col_map):
                    c = _cl(col)
                    _formula(ws, row, col,
                             f"={c}{seg_oi_row}/{c}{seg_rev_row}",
                             fmt=PCT_FMT)
                row += 1

    else:
        _label(ws, row, "Segment data not available in filings — populate manually")
        row += 1

    ws.freeze_panes = "C6"


# ---------------------------------------------------------------------------
# Verification checks
# ---------------------------------------------------------------------------

def _run_checks(data, col_map, annual_years):
    """Run verification checks. Returns list of (check_name, status, detail)."""
    checks = []

    bs_data = data.get("balance_sheet", {})
    bs_items = bs_data.get("line_items", [])

    # BS balance check: Total Assets = Total Liabilities + Equity
    total_assets_item = None
    total_liab_equity_item = None
    for item in bs_items:
        lbl_lower = item.get("label", "").lower()
        if "total assets" in lbl_lower and item.get("type") == "subtotal":
            total_assets_item = item
        if ("total liabilities" in lbl_lower and "equity" in lbl_lower
                and item.get("type") == "subtotal"):
            total_liab_equity_item = item

    if total_assets_item and total_liab_equity_item:
        for yr in annual_years:
            ta_val = None
            tle_val = None
            for pk in total_assets_item.get("values", {}):
                if str(yr) in pk:
                    ta_val = total_assets_item["values"][pk]
            for pk in total_liab_equity_item.get("values", {}):
                if str(yr) in pk:
                    tle_val = total_liab_equity_item["values"][pk]
            if ta_val is not None and tle_val is not None:
                diff = abs(ta_val - tle_val)
                status = "PASS" if diff < 1 else "FAIL"
                checks.append((
                    f"BS Balance FY{yr}",
                    status,
                    f"Assets={ta_val:,.0f}, L+E={tle_val:,.0f}, diff={diff:,.0f}",
                ))

    # IS subtotal check: do reported subtotals match component sums?
    is_data = data.get("income_statement", {})
    is_items = is_data.get("line_items", [])
    # This check would require re-computing formulas — skip for now,
    # the Excel formulas will show discrepancies directly

    checks.append(("IS Formulas", "INFO",
                    f"{len(is_items)} line items with Excel formulas for subtotals"))

    # CF check placeholder
    cf_data = data.get("cash_flow", {})
    cf_items = cf_data.get("line_items", [])
    checks.append(("CF Structure", "INFO",
                    f"{len(cf_items)} line items"))

    return checks


def _write_checks_tab(wb, checks):
    """Write verification checks to a Checks tab."""
    ws = wb.create_sheet(title="Checks")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 60

    ws.cell(row=1, column=2, value="Check").font = HEADER_FONT
    ws.cell(row=1, column=3, value="Status").font = HEADER_FONT
    ws.cell(row=1, column=4, value="Detail").font = HEADER_FONT

    for i, (name, status, detail) in enumerate(checks, start=2):
        ws.cell(row=i, column=2, value=name)
        cell = ws.cell(row=i, column=3, value=status)
        if status == "PASS":
            cell.font = Font(color="008000", bold=True)
        elif status == "FAIL":
            cell.font = Font(color="FF0000", bold=True)
        else:
            cell.font = Font(color="808080")
        ws.cell(row=i, column=4, value=detail)


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------

def build_model(ticker):
    """Build the financial model Excel workbook for a ticker."""
    # Load data
    data_path = DATA_DIR / ticker / "financials" / "extracted_data.json"
    if not data_path.exists():
        print(f"Error: no extracted data at {data_path}")
        print(f"  Run: python main.py extract {ticker}")
        return

    with open(data_path) as f:
        data = json.load(f)

    with open(CONFIG_PATH) as f:
        companies = json.load(f)
    company_name = companies.get(ticker, {}).get("name", ticker)

    # Classify periods from IS data
    is_pks = data.get("income_statement", {}).get("period_keys", [])
    annual_years, quarterly = _classify_is_periods(is_pks)

    if not annual_years:
        print(f"No annual periods found for '{ticker}'.")
        return

    col_map = _build_col_map(annual_years, quarterly)

    wb = Workbook()

    # --- Tab 1: FS ---
    fs_name = f"{company_name} FS"[:31]
    ws = wb.active
    ws.title = fs_name

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 45
    for col in _all_data_cols(col_map):
        ws.column_dimensions[_cl(col)].width = 14
    ws.column_dimensions[_cl(col_map["spacer_col"])].width = 2

    # Company name
    cell = ws.cell(row=3, column=2, value=company_name)
    cell.font = Font(bold=True)

    # Income Statement
    next_row, revenue_row, label_to_row = _write_is_section(
        ws, data, col_map, annual_years, quarterly
    )

    # Y/Y growth
    next_row = _write_yy_section(ws, next_row, col_map, annual_years, label_to_row)

    # Q/Q growth
    next_row = _write_qq_section(ws, next_row, col_map, annual_years, label_to_row)

    # % of sales
    next_row = _write_pct_section(ws, next_row, col_map, annual_years,
                                   label_to_row, revenue_row)

    # Balance Sheet
    bs_result = _write_bs_section(ws, data, col_map, annual_years, next_row)
    if isinstance(bs_result, tuple):
        next_row, label_to_row_bs = bs_result
    else:
        next_row = bs_result

    # Cash Flow
    next_row = _write_cf_section(ws, data, col_map, annual_years, next_row)

    # Freeze panes
    ws.freeze_panes = "C6"

    # --- Tab 2: Drivers ---
    _build_drivers_sheet(wb, company_name, data, col_map, annual_years,
                         fs_name, revenue_row)

    # --- Checks ---
    checks = _run_checks(data, col_map, annual_years)
    _write_checks_tab(wb, checks)

    # Print checks
    print("\n=== Verification Checks ===")
    for name, status, detail in checks:
        print(f"  [{status}] {name}: {detail}")

    # Save
    MODELS_DIR.mkdir(parents=True, exist_ok=True)
    output_path = MODELS_DIR / f"{ticker}_model.xlsx"
    wb.save(output_path)

    n_annual = len(annual_years)
    n_quarterly = len(col_map["ordered_q"])
    n_is = len(data.get("income_statement", {}).get("line_items", []))
    n_bs = len(data.get("balance_sheet", {}).get("line_items", []))
    n_cf = len(data.get("cash_flow", {}).get("line_items", []))
    print(f"\nModel saved to {output_path}")
    print(f"  {n_annual} annual periods, {n_quarterly} quarterly periods")
    print(f"  IS: {n_is} line items, BS: {n_bs} line items, CF: {n_cf} line items")
